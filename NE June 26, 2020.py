from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup


path="E:\\Selenium Bizmatics Inc\\Unprocessed automation.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
row=sheet.max_row
cols=sheet.max_column

driver = webdriver.Chrome(executable_path="C:\ChromeSeleniumDriver\chromedriver.exe")

driver.get("https://bellairefc.prognocis.com/prognocis/scrUserLogin.jsp?clinic=bellairefc")
username=driver.find_element_by_id('username')
username.is_enabled()
driver.find_element_by_id('username').send_keys('Kraig')

password=driver.find_element_by_id('password')
password.is_enabled()
driver.find_element_by_id('password').send_keys('F3rSmartview')

login=driver.find_element_by_id('login')
login.is_enabled()
driver.find_element_by_id('login').click()
time.sleep(2)


for r in range(2,row-1):
    status=sheet.cell(row=r,column=13).value
    if status is not None:
        continue

    CalimNumber=sheet.cell(row=r,column=1).value
    print("Row Number:-", r)
    handles = driver.window_handles
    driver.switch_to.window(handles[1])
    driver.switch_to.frame(0)
    driver.find_element_by_id("menuClaims").click()
    driver.switch_to.default_content()
    driver.find_element_by_link_text("Edit Claims").click()
    handles = driver.window_handles
    driver.switch_to.window(handles[2])

    driver.find_element_by_id("BLH_CLAIM_ID").click()
    driver.find_element_by_id("BLH_CLAIM_ID").clear()
    driver.find_element_by_id("BLH_CLAIM_ID").send_keys(CalimNumber)
    time.sleep(4)

    try:
        NoRecordFound=driver.find_element_by_xpath("//table[@id='table']/tbody/tr[3]/td/div/table/tbody/div/p").text
        if NoRecordFound=="No Record Found":
            driver.find_element_by_id("close").click()
            sheet.cell(row=r, column=13).value = 'No Record Found'
            continue
    except:
        pass

    driver.find_element_by_xpath("//tr[@id='rownum0']/td[3]/span").click()
    time.sleep(7)
    try:
        driver.switch_to.window(handles[1])
        #time.sleep(2)
        driver.switch_to.frame(4)
        #time.sleep(2)
        driver.switch_to.frame(1)
        time.sleep(2)
        driver.find_element_by_id("ok").click()
    except:
        pass

    time.sleep(2)
    driver.switch_to.window(handles[1])
    #time.sleep(2)
    driver.switch_to.frame('mainFrame')
    time.sleep(2)
    patins=driver.find_element_by_id("patins")
    patins.is_enabled()
    time.sleep(2)
    driver.find_element_by_id("patins").click()

    time.sleep(2)
    driver.switch_to.default_content()
    #time.sleep(2)
    driver.switch_to.frame('jqmContent')
    #time.sleep(2)
    driver.switch_to.frame('mainFrame')
    #time.sleep(4)

    try:
        element = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.ID, "eligibilitycheck"))
        )
    except:
        pass

    driver.find_element_by_id("eligibilitycheck").click()
    time.sleep(8)


    try:
        driver.find_element_by_id("subeligibility").click()
    except:

        time.sleep(5)
        driver.find_element_by_id("subeligibility").click()

    time.sleep(2)
    driver.switch_to.default_content()  #parent
    #time.sleep(2)
    driver.switch_to.default_content()  #parent
    print('.....parent...........')
    #time.sleep(3)
    driver.switch_to.frame(5)
    #time.sleep(1)
    driver.switch_to.frame(0)
    #time.sleep(1)
    driver.switch_to.frame(1)
    #time.sleep(3)
    try:
        Eligibility=driver.find_element_by_id("trnEligibilityStatus").text
    except:
        Eligibility = "Error "

    #print(Eligibility)
    time.sleep(3)
    if Eligibility != "Active Coverage":
        driver.find_element_by_id("cancel").click()
        driver.switch_to.default_content()
        driver.switch_to.frame(4)
        time.sleep(2)
        driver.switch_to.frame(1)
        time.sleep(2)
        driver.find_element_by_id("cancel").click()
        driver.switch_to.default_content()
        sheet.cell(row=r, column=13).value = 'Not Processed'
        workbook.save(path)
    else:
        driver.find_element_by_id("cancel").click()
        #time.sleep(2)
        driver.switch_to.default_content()
        #time.sleep(2)
        driver.switch_to.frame(4)
        #time.sleep(2)
        driver.switch_to.frame(1)
        #time.sleep(2)
        driver.find_element_by_id("cancel").click()
        time.sleep(3)
        try:
            driver.switch_to.alert.accept()
            #driver.switch_to.default_content()
        except:
            pass
        time.sleep(3)
        driver.switch_to.default_content()
        #time.sleep(1)
        driver.switch_to.frame(1)
        #time.sleep(1)
        driver.find_element_by_id("msBlh_pos_code").click()
        time.sleep(1)
        Select(driver.find_element_by_id("msBlh_pos_code")).select_by_visible_text("20 - Urgentcare")
        time.sleep(1)
        driver.find_element_by_id("msBlh_pos_code").click()
        time.sleep(1)
        driver.find_element_by_id("msBlh_Business_Unit").click()
        time.sleep(1)
        Select(driver.find_element_by_id("msBlh_Business_Unit")).select_by_visible_text(
            "UC - Northeast Urgent Care Corporation")
        time.sleep(1)
        driver.find_element_by_id("msBlh_Business_Unit").click()
        time.sleep(1)

        for addi in range(0,13):
            image = str('addicdhcpc' + str(addi))
            RowDisplayed = driver.find_element_by_id(image).is_displayed()
            if RowDisplayed is False:
                Scroldown = driver.find_element_by_id(image)
                driver.execute_script("arguments[0].scrollIntoView();", Scroldown)
                RowDisplayed = driver.find_element_by_id(image).is_displayed()
                if RowDisplayed is False:
                    break
                else:
                    driver.find_element_by_id(image).click()
                    driver.switch_to.default_content()
                    driver.switch_to.frame('jqmContent')
            else:
                driver.find_element_by_id(image).click()
                driver.switch_to.default_content()
                # time.sleep(2)
                driver.switch_to.frame('jqmContent')
            try:
                time.sleep(3)
                checkStatus0 = driver.find_element_by_id("chk0").is_selected()
                if checkStatus0 is False:
                    driver.find_element_by_id("chk0").click()

                checkStatus1 = driver.find_element_by_id("chk1").is_selected()
                if checkStatus1 is False:
                    driver.find_element_by_id("chk1").click()

                checkStatus2 = driver.find_element_by_id("chk2").is_selected()
                if checkStatus2 is False:
                    driver.find_element_by_id("chk2").click()

                checkStatus3 = driver.find_element_by_id("chk3").is_selected()
                if checkStatus3 is False:
                    driver.find_element_by_id("chk3").click()

                driver.switch_to.default_content()
                driver.switch_to.frame('jqmContent')
                driver.find_element_by_id("ok").click()
                driver.switch_to.default_content()  # parent
                driver.switch_to.frame(1)
            except:
                driver.switch_to.default_content()
                # time.sleep(2)
                driver.switch_to.frame('jqmContent')
                # time.sleep(2)
                driver.find_element_by_id("ok").click()
                time.sleep(1)
                driver.switch_to.default_content()  # parent
                driver.switch_to.frame(1)

        time.sleep(1)
        driver.switch_to.default_content()
        driver.switch_to.frame(1)
        time.sleep(1)
        driver.find_element_by_id('save').click()
        time.sleep(1)

        try:  # click on pop messsage
            driver.switch_to.default_content()
            time.sleep(1)
            driver.switch_to.frame(4)
            time.sleep(1)
            driver.find_element_by_name("ok").click()
        except:
            pass

        sheet.cell(row=r, column=13).value = 'Saved'
        workbook.save(path)
        driver.switch_to.default_content()
